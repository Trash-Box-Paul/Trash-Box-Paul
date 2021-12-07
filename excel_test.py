import bs4
import openpyxl
import time
from datetime import datetime
import webdrivermanager
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import win32api, win32con
import sys
from bs4 import BeautifulSoup
from selenium.webdriver.common.keys import Keys
from debug_browser import DebugBrowser


class TakeTasks:

    def __init__(self):
        self.driver = None
        self.driver_setup()
        self.spread_sheet_read = openpyxl.load_workbook('Latest_Spread_Sheet.xlsm', data_only=True)
        self.spread_sheet_write = openpyxl.load_workbook('Latest_Spread_Sheet.xlsm', data_only=False)
        self.status_sheet_read = self.spread_sheet_read.get_sheet_by_name('general')
        self.in_progress_sheet_read = self.spread_sheet_read.get_sheet_by_name('in progress')
        self.pend_sheet_read = self.spread_sheet_read.get_sheet_by_name('pending')
        self.in_progress_sheet_write = self.spread_sheet_write.get_sheet_by_name('in progress')
        self.pend_sheet_write = self.spread_sheet_write.get_sheet_by_name('pending')
        # ------------------------------------------------------------------------------------------------------
        self.info_collect_list_read = openpyxl.load_workbook('pytest.xlsx', data_only=True)
        self.info_collect_list_write = openpyxl.load_workbook('pytest.xlsx', data_only=False)
        self.testsheet = self.info_collect_list_write.get_sheet_by_name('Sheet1')
        self.cloudsheet = self.info_collect_list_read.get_sheet_by_name('cloud')
        self.newtp_read_sheet = self.info_collect_list_read.get_sheet_by_name('newtp')
        self.newtp_write_sheet = self.info_collect_list_write.get_sheet_by_name('newtp')
        self.count = 0

    def driver_setup(self):
        print("22222222222222")
        chrome_driver = webdrivermanager.ChromeDriverManager()
        try:
            print("1111111111111111")
            self.driver = webdriver.Chrome(executable_path=chrome_driver.get_driver_filename(),
                                           options=DebugBrowser().debug_chrome())
            print("????????????????")
        except:
            chrome_driver.download_and_install(chrome_driver.get_latest_version())
            time.sleep(3)
            print("!!!!!!!!!!!!!!!!")
            time.sleep(3)
            self.driver = webdriver.Chrome(executable_path=chrome_driver.get_driver_filename(),
                                           options=DebugBrowser().debug_chrome())

    def grab_task_name_ID(self):
        # all_handle = self.driver.window_handles  # get all handles
        target_url = "https://907826.app.netsuite.com/app/center/card.nl?sc=-29&whence="
        # self.driver.switch_to.window(all_handle[-1])  # Switch to the new pop-up window
        # # 2 | open | /app/center/card.nl?sc=-29&whence= |\
        # # time.sleep(2)
        self.driver.get(target_url)
        # if not ("https://907826.app.netsuite.com/app/center/" in self.driver.current_url):
        #     win32api.MessageBox(0, "Please login first and try again. :)", "Please Login",
        #                         win32con.MB_OK)
        #     return 'Mission Failed'
        table_content = "/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div[1]/div[2]/div/div/div/div"
        number_sum = "/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div[1]/div[2]/div/div/form/div[" \
                     "2]/table/tbody/tr/td/table/tbody/tr/td/a "
        ele = self.wait(number_sum)
        html = ele.get_attribute('innerHTML')
        case_sum = int(html)
        time.sleep(1)
        ele = self.wait(table_content)
        html = ele.get_attribute('innerHTML')
        soup = BeautifulSoup(html, 'html5lib')
        tables = soup.findAll('table')
        tab = tables[0]
        table_body = tab.tbody
        tr_group = table_body.find_all('tr')
        target = int(len(tr_group) - 1)  # number of tr
        task_name = []
        task_id = []
        task_customer = []
        print(target)
        for tr in tr_group:
            if not ("text" in tr['class']):
                td_group = tr.find_all('td')
                task_name.append(td_group[2].text)
                task_id.append(str(td_group[1].find_all('a')[1]['href']).split('=')[1])
                task_customer.append(td_group[8].find('a').text)
                # .get_attribute("href").split('=')[1]
        print(task_id)
        print(task_name)
        print(task_customer)
        for num in range(0, case_sum):
            self.pend_sheet_write['A' + str(num + 2)] = task_name[num].strip()
            self.pend_sheet_write['B' + str(num + 2)] = task_id[num].strip()
            self.pend_sheet_write['C' + str(num + 2)] = task_customer[num].split('[')[0].strip()
        self.spread_sheet_write.save('Latest_Spread_Sheet(after).xlsx')

    def edit_note(self, task_num, new_note):
        edit_icon = "/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div[1]/div[2]/div/div/div/div/table/tbody/tr["
        edit_icon_ex = "]/td[2]/a[1]"
        table_content = "/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div[1]/div[2]/div/div/div/div"
        time.sleep(2)
        ele = self.wait(edit_icon + str(task_num) + edit_icon_ex)
        self.driver.get(ele.get_attribute("href"))
        text_box = "/html/body/div[1]/div[2]/div[3]/table[1]/tbody/tr[3]/td/div[" \
                   "1]/div/div/form/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/div/span[2]/span/textarea"
        # time.sleep(1)
        ele = self.wait(text_box)
        current_handle = self.driver.current_window_handle
        if ele.text is not None and ele.text.strip() == new_note.strip():
            self.driver.get("https://907826.app.netsuite.com/app/center/card.nl?sc=-29&whence=")
        else:
            self.driver.execute_script('arguments[0].click()', ele)
            ele.send_keys(Keys.CONTROL + 'a')
            ele.send_keys(new_note)
            self.driver.execute_script('arguments[0].click()', self.wait("/html/body/div[1]/div[2]/div["
                                                                         "3]/form/table/tbody/tr["
                                                                         "2]/td/table/tbody/tr[4]/td[ "
                                                                         "2]/table/tbody/tr[8]/td/div/span["
                                                                         "2]/span/input"))
            # print(new_note)
            # time.sleep(10)
            ele = self.wait("/html/body/div[1]/div[2]/div[3]/form/table/tbody/tr["
                            "1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td[2]/input")
            while "&e=T" in self.driver.current_url:
                try:
                    self.driver.execute_script(ele.get_attribute('onclick'))
                    self.driver.get("https://907826.app.netsuite.com/app/center/card.nl?sc=-29&whence=")
                except:
                    self.driver.get("https://907826.app.netsuite.com/app/center/card.nl?sc=-29&whence=")

    def grab_note(self, task_num, new_note):
        edit_icon = "/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div[1]/div[2]/div/div/div/div/table/tbody/tr["
        edit_icon_ex = "]/td[2]/a[1]"
        table_content = "/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div[1]/div[2]/div/div/div/div"
        time.sleep(2)
        ele = self.wait(edit_icon + str(task_num) + edit_icon_ex)
        self.driver.get(ele.get_attribute("href"))
        text_box = "/html/body/div[1]/div[2]/div[3]/table[1]/tbody/tr[3]/td/div[" \
                   "1]/div/div/form/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/div/span[2]/span/textarea"
        # time.sleep(1)
        ele = self.wait(text_box)
        current_handle = self.driver.current_window_handle
        if ele.text is not None and ele.text.strip() == new_note.strip():
            self.driver.get("https://907826.app.netsuite.com/app/center/card.nl?sc=-29&whence=")
        else:
            self.driver.execute_script('arguments[0].click()', ele)
            ele.send_keys(Keys.CONTROL + 'a')
            ele.send_keys(new_note)
            self.driver.execute_script('arguments[0].click()', self.wait("/html/body/div[1]/div[2]/div["
                                                                         "3]/form/table/tbody/tr["
                                                                         "2]/td/table/tbody/tr[4]/td[ "
                                                                         "2]/table/tbody/tr[8]/td/div/span["
                                                                         "2]/span/input"))
            # print(new_note)
            # time.sleep(10)
            ele = self.wait("/html/body/div[1]/div[2]/div[3]/form/table/tbody/tr["
                            "1]/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td[2]/input")
            while "&e=T" in self.driver.current_url:
                try:
                    self.driver.execute_script(ele.get_attribute('onclick'))
                    self.driver.get("https://907826.app.netsuite.com/app/center/card.nl?sc=-29&whence=")
                except:
                    self.driver.get("https://907826.app.netsuite.com/app/center/card.nl?sc=-29&whence=")

    def update_all_notes(self):
        cur_handle = self.driver.current_window_handle  # get current handle
        all_handle = self.driver.window_handles  # get all handles
        target_url = "https://907826.app.netsuite.com/app/center/card.nl?sc=-29&whence="
        self.driver.switch_to.window(all_handle[-1])  # Switch to the new pop-up window
        # 2 | open | /app/center/card.nl?sc=-29&whence= |\
        # time.sleep(2)
        self.driver.get(target_url)
        if not ("https://907826.app.netsuite.com/app/center/" in self.driver.current_url):
            win32api.MessageBox(0, "Please login first and try again. :)", "Please Login",
                                win32con.MB_OK)
            return 'Mission Failed'
        table_content = "/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div[1]/div[" \
                        "2]/div/div/div/div"
        number_sum = "/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div[1]/div[2]/div/div/form/div[" \
                     "2]/table/tbody/tr/td/table/tbody/tr/td/a "
        ele = self.wait(number_sum)
        html = ele.get_attribute('innerHTML')
        case_sum = int(html)
        excel_sum = int(self.status_sheet_read['B1'].value) - 1
        ele = self.wait(table_content)
        html = ele.get_attribute('innerHTML')
        soup = BeautifulSoup(html, 'html5lib')
        tables = soup.findAll('table')
        tab = tables[0]
        table_body = tab.tbody
        tr_group = table_body.find_all('tr')
        target = int(len(tr_group) - 1)  # number of tr
        task_name = []
        print(target)
        for tr in tr_group:
            if not ("text" in tr['class']):
                td_group = tr.find_all('td')
                task_name.append(td_group[2].text)
        print(task_name)
        for num in range(0, case_sum):
            for excel_num in range(0, excel_sum):
                if str.strip(self.in_progress_sheet_read['A' + str(excel_num + 2)].value) == str.strip(task_name[num]):
                    if self.in_progress_sheet_read['E' + str(excel_num + 2)].value is not None:
                        self.edit_note(num + 1, self.in_progress_sheet_read['E' + str(excel_num + 2)].value)
                        break

    def grab_pend_task(self):
        target_url = "https://907826.app.netsuite.com/app/center/card.nl?sc=-29&whence="
        # cur_handle = self.driver.current_window_handle  # get current handle
        # all_handle = self.driver.window_handles  # get all handles
        # for h in all_handle:
        #     if h != cur_handle:
        #         self.driver.switch_to.window(h)  # Switch to the new pop-up window
        #         break
        # 2 | open | /app/center/card.nl?sc=-29&whence= |\
        # time.sleep(2)
        self.driver.get(target_url)
        self.switch_status("pending")
        time.sleep(2)
        self.switch_status("progressing")
        time.sleep(2)
        self.switch_status("pending")

    def switch_status(self, status):
        js_top = "var q=document.documentElement.scrollTop=0"
        self.driver.execute_script(js_top)
        tab_case = "/html/body/div[1]/div[1]/div[2]/ul[3]/li/a"
        self.driver.execute_script('arguments[0].click()', self.wait(tab_case))
        title = "/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div[1]/div[1]/h2"
        element = self.wait(title)
        actions = ActionChains(self.driver)
        actions.move_to_element(element).perform()
        # 1 | mouseMoveAt | Title: Paul's All case view | hover element
        # element.click()
        element = self.wait("/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div[1]/div[1]/div/div")
        actions = ActionChains(self.driver)
        actions.move_to_element(element).perform()
        self.driver.execute_script('arguments[0].click()', element.click)
        # 2 | mouseMoveAt | Configure Icon | hover element
        element = self.wait("/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div[1]/div[1]/div/div/ul/li[3]/a")
        while element.get_attribute("innerHTML") is None:
            element = self.driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div["
                                                         "1]/div[1]/div/div/ul/li[3]/a")
        while not ("Edit" in element.get_attribute("innerHTML")):
            element = self.driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div["
                                                         "1]/div[1]/div/div/ul/li[3]/a")
        actions = ActionChains(self.driver)
        actions.move_to_element(element).perform()
        self.driver.execute_script('arguments[0].click()', element)
        # 3 | mouseMoveAt and click | Edit Icon | hover element
        self.driver.execute_script('arguments[0].click()', self.wait("/html/body/div[1]/div[2]/div[3]/table["
                                                                     "1]/tbody/tr[1]/td/table/tbody/tr/td[2]/a"))
        criteria_subject = "/html/body/div[1]/div[2]/div[3]/table[1]/tbody/tr[3]/td/div[1]/div/div/table/tbody/tr[" \
                           "2]/td/div/div[8]/div/form/div[6]/table/tbody/tr[4]/td[1] "
        self.driver.execute_script('arguments[0].click()', self.wait(criteria_subject))
        # actions = ActionChains(self.driver)
        # actions.move_to_element(criteria_subject).perform()
        arrow = "/html/body/div[1]/div[2]/div[3]/table[1]/tbody/tr[3]/td/div[1]/div/div/table/tbody/tr[2]/td/div/div[" \
                "8]/div/form/div[6]/table/tbody/tr[4]/td[1]/div/div/span/span[2]/a"

        element = self.wait(arrow)
        actions = ActionChains(self.driver)
        actions.move_to_element(element).perform()
        self.driver.execute_script('arguments[0].click()', element)
        iframe = "/html/body/div[9]/div[2]/div[1]/div/div/iframe"
        element = self.wait(iframe)
        self.driver.switch_to.frame(element)
        pending_line = "/html/body/div[1]/div/div[4]/form/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[" \
                       "2]/td/div/span[2]/span/div[2]/table/tbody/tr[2]/td/a"
        progress_line = "/html/body/div[1]/div/div[4]/form/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[" \
                        "2]/td/div/span[2]/span/div[2]/table/tbody/tr[3]/td/a "
        if status == "pending":
            self.driver.execute_script('arguments[0].click()', self.wait(pending_line))
            # 4 | Input | Search Key Words
        else:
            self.driver.execute_script('arguments[0].click()', self.wait(progress_line))
        # 4 | Input | Search Key Words
        temp_element = "/html/body/div[1]/div/div[4]/form/table/tbody/tr[1]/td/table/tbody/tr/td/table/tbody/tr/td[" \
                       "1]/table/tbody/tr/td[2]/input"
        self.driver.execute_script('arguments[0].click()', self.wait(temp_element))
        # 5 | Click | Set
        temp_element = "/html/body/div[1]/div[2]/div[3]/table[2]/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td[" \
                       "2]/input "
        self.driver.execute_script('arguments[0].click()', self.wait(temp_element))
        # 6 | Click | Save

    def cloud_ftp(self, profile_name):
        target_url = "http://psdtool.tc.net/psdTool/"
        cur_handle = self.driver.current_window_handle  # get current handle
        all_handle = self.driver.window_handles  # get all handles
        for h in all_handle:
            if h != cur_handle:
                self.driver.switch_to.window(h)  # Switch to the new pop-up window
                break
        self.driver.get(target_url)
        search_input = "/html/body/form/table/tbody/tr[2]/td[2]/div/table/tbody/tr/td[2]/input"
        ele = self.wait(search_input)
        ele.send_keys(profile_name)
        self.driver.find_element(By.XPATH, "/html/body/form/table/tbody/tr[2]"
                                           "/td[2]/div/table/tbody/tr/td[3]/input").click()

        psd_title = "/html/body/form/table/tbody/tr[3]/td/div/div[1]/div/div[1]"
        if "ProfileDetails" not in self.driver.current_url:
            ele = self.wait("/html/body/form/table/tbody/tr[3]/td/div/div[2]/div/div/div/div/table/tbody/tr[2]/td[2]/a")
            time.sleep(2)
            self.driver.execute_script('arguments[0].click()', ele)
        notes_input = "/html/body/form/table/tbody/tr[3]/td/div/table/tbody/tr[3]/td/table/tbody/tr[2]/td/div[" \
                      "1]/div[2]/div/table/tbody/tr[2]/td[2]/table/tbody/tr/td[2]/textarea"
        try:
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, notes_input))
            )
        finally:
            time.sleep(1)
            qualifier = self.driver.find_element(By.XPATH, "/html/body/form/table/tbody/tr[3]/td/div/table/tbody/tr["
                                                           "3]/td/table/tbody/tr[2]/td/div[1]/div[4]/table/tbody/tr["
                                                           "2]/td[1]").get_attribute("innerHTML")
            ediid = self.driver.find_element(By.XPATH, "/html/body/form/table/tbody/tr[3]/td/div/table/tbody/tr["
                                                       "3]/td/table/tbody/tr[2]/td/div[1]/div[4]/table/tbody/tr[2]/td["
                                                       "2]").get_attribute("innerHTML")
            ele = self.driver.find_element(By.XPATH, notes_input)
            username = qualifier + ediid
            print(username)
            ele.send_keys("\n" + "Cloud SFTP:" + "\n" + "U: " + username + "\n")
            var = profile_name.split()
            profile_id = self.driver.find_element(By.XPATH, "/html/body/form/table/tbody/tr[3]/td/div/table/tbody/tr["
                                                            "3]/td/table/tbody/tr[2]/td/div[1]/div[1]/table/tbody/tr["
                                                            "1]/td[2]").get_attribute("innerHTML")
            password = var[0] + var[1] + profile_id + "!"
            print(password)
            ele.send_keys("P: " + password)
            self.driver.execute_script('arguments[0].click()', self.wait("/html/body/form/table/tbody/tr["
                                                                         "3]/td/div/table/tbody/tr["
                                                                         "3]/td/table/tbody/tr[2]/td/div[1]/div["
                                                                         "2]/div/table/tbody/tr[3]/td/input"))
            time.sleep(2)
            profile_manage = "/html/body/form/table/tbody/tr[2]/td[1]/table/tbody/tr/td/table/tbody/tr/td[" \
                             "4]/table/tbody/tr/td[1]/a"
            ftp_setup = "/html/body/form/table/tbody/tr[2]/td[1]/table/tbody/tr/td/div[4]/table/tbody/tr[" \
                        "6]/td/table/tbody/tr/td "
            setup_inbox = "/html/body/form/table/tbody/tr[3]/td/div/table/tbody/tr[2]/td[1]/input[1]"
            element = self.driver.find_element(By.XPATH, profile_manage)
            actions = ActionChains(self.driver)
            actions.move_to_element(element).perform()

            ele = self.wait(ftp_setup)
            actions = ActionChains(self.driver)
            actions.move_to_element(ele)
            actions.click(ele).perform()
            try:
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, setup_inbox))
                )
            finally:
                ele = self.driver.find_element(By.XPATH, setup_inbox)
                ele.send_keys(profile_name)
                self.driver.execute_script('arguments[0].click()', self.wait("/html/body/form/table/tbody/tr["
                                                                             "3]/td/div/table/tbody/tr[2]/td[ "
                                                                             "1]/input[2]"))
                self.driver.find_element(By.XPATH, "/html/body/form/table/tbody/tr[3]/td/div/div["
                                                   "1]/div/table/tbody/tr[3]/td[2]/input").send_keys(username)
                self.driver.find_element(By.XPATH, "/html/body/form/table/tbody/tr[3]/td/div/div["
                                                   "1]/div/table/tbody/tr[4]/td[2]/input").send_keys(password)
                self.driver.execute_script('arguments[0].click()', self.wait("/html/body/form/table/tbody/tr["
                                                                             "3]/td/div/div[1]/div/table/tbody/tr["
                                                                             "5]/td/input[2]"))
                self.driver.execute_script('arguments[0].click()', self.wait("/html/body/form/table/tbody/tr["
                                                                             "3]/td/div/div[1]/div/table/tbody/tr["
                                                                             "6]/td/input"))
                time.sleep(2)

    def set_all_cloud(self):
        for item in list(self.cloudsheet.columns)[0]:
            self.cloud_ftp(str(item.value))

    def send_initial_emails(self, email, tp_name):
        self.driver.get("https://907826.app.netsuite.com/app/center/card.nl?sc=-29&whence=")
        table_content = "/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div[1]/div[2]/div/div/div/div"
        ele = self.wait(table_content)
        html = ele.get_attribute('innerHTML')
        soup = BeautifulSoup(html, 'html5lib')
        tab = soup.findAll('table')
        tr_group = tab[0].tbody.find_all('tr')
        count = 0
        for tr in tr_group:
            count += 1
            print(tr.find_all('td')[2].text)
            if tp_name in tr.find_all('td')[2].text:
                project_path = "/html/body/div[1]/div[2]/div/div/div/div[5]/div[2]/div[1]/div[" \
                               "2]/div/div/div/div/table/tbody/tr[" + str(count) + "]/td[7]/a"
                ele = self.wait(project_path)
                cust_name = ele.text.split('[')[0]
                self.driver.get(ele.get_property('href'))
                self.driver.execute_script('arguments[0].click()', self.wait_id("custom100txt"))
                self.driver.execute_script('arguments[0].click()', self.wait_id("newmessage"))
                # time.sleep(10)
                all_handle = self.driver.window_handles  # get all handle
                for h in all_handle:
                    self.driver.switch_to.window(h)  # Switch to the new pop-up window
                    if "crm/common/crmmessage" in self.driver.current_url:
                        break
                print(email[0])
                current_handle = self.driver.current_window_handle
                self.wait_id("recipientemail").send_keys(email[0])
                if len(email) > 1:
                    self.driver.execute_script('arguments[0].click()', self.wait("/html/body/div[1]/div/div[4]/table["
                                                                                 "1]/tbody/tr[3]/td/div["
                                                                                 "1]/div/div/table/tbody/tr[ "
                                                                                 "2]/td/div/div[9]/div/form/div["
                                                                                 "6]/table/tbody/tr[2]/td[2]/div"))
                    self.wait_id("email").send_keys(email[1])
                    self.driver.execute_script('arguments[0].click()', self.wait_id("otherrecipientslist_addedit"))
                self.driver.execute_script('arguments[0].click()', self.wait_id("messagestxt"))
                self.wait_id("template_display").send_keys("Paul")
                self.wait_id("template_display").send_keys(Keys.ENTER)
                time.sleep(2)
                self.wait_id('subject').send_keys(' ' + tp_name)
                ele = self.wait("/html/body/div[1]/div/div[4]/table[1]/tbody/tr[3]/td/div["
                                "2]/div/div/form/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[6]/td/div/span["
                                "2]/div/div/div/iframe")
                self.driver.switch_to.frame(ele)
                self.driver.execute_script('arguments[0].click()', self.wait("/html/body"))
                self.wait("/html/body").send_keys(Keys.CONTROL + Keys.HOME)
                self.wait("/html/body").send_keys(
                    "Good morning " + tp_name + "," + "\n" + "\n" + cust_name + "has selected "
                                                                                "TrueCommerce EDI "
                                                                                "Solutions Group to be "
                                                                                "their EDI software "
                                                                                "Service Provider. In "
                                                                                "order for us to "
                                                                                "complete this "
                                                                                "relationship setup, "
                                                                                "please "
                                                                                "provide the following "
                                                                                "information:")
                time.sleep(3)
                self.driver.switch_to.window(current_handle)
                self.driver.execute_script('arguments[0].click()', self.wait("/html/body/div[1]/div/div["
                                                                             "4]/form/table/tbody/tr["
                                                                             "1]/td/table/tbody/tr/td/table/tbody/tr"
                                                                             "/td[1]/table/tbody/tr/td[2]/input"))
                break

    def send_all_tps(self):
        print(1)
        total_count = 1
        for item in list(self.newtp_read_sheet.columns)[0]:
            if item is not None and item.value is not None:
                self.new_window("https://907826.app.netsuite.com/app/center/card.nl?sc=-29&whence=")
                ele = self.wait("/html/body/div[1]/div[1]/div[1]/div[4]/input[1]")
                search_name = str(self.newtp_read_sheet['B' + str(total_count)].value)
                print(search_name)
                ele.send_keys("part: " + search_name)
                ele.send_keys(Keys.ENTER)
                time.sleep(3)
                if 'globalsearch' in str(self.driver.current_url):
                    self.newtp_write_sheet['C' + str(total_count)] = "Multiple Records"
                    total_count += 1
                    self.driver.close()
                    continue
                ele = self.wait_id('s_relationtxt')  # Relationships
                time.sleep(3)
                self.driver.execute_script('arguments[0].click()', ele)
                print(1)
                time.sleep(3)
                try:
                    ele = self.driver.find_element(By.ID, "contact__div")
                    html = ele.get_attribute('innerHTML')
                    soup = BeautifulSoup(html, 'html5lib')
                    tables = soup.findAll('table')
                    tab = tables[0]
                    email_list = []
                    table_body = tab.tbody
                    tr_group = table_body.find_all('tr')
                    td_group = tr_group[0].find_all('td')
                    if td_group[0].text == "No records to show.":
                        self.newtp_write_sheet['C' + str(total_count)] = "No Contacts"
                        total_count += 1
                        self.driver.close()
                        continue
                    print(str(td_group[5].find_all('a')[1].text))
                    self.newtp_write_sheet['C' + str(total_count)] = td_group[5].find_all('a')[1].text
                    email_list.append(td_group[5].find_all('a')[1].text)
                    if len(tr_group) > 1:
                        td_group = tr_group[1].find_all('td')
                        self.newtp_write_sheet['D' + str(total_count)] = td_group[5].find_all('a')[1].text
                        print(str(td_group[5].find_all('a')[1].text))
                        email_list.append(td_group[5].find_all('a')[1].text)
                    current_handle = self.driver.current_window_handle
                    self.send_initial_emails(email_list, search_name)
                    self.driver.switch_to.window(current_handle)
                    self.newtp_write_sheet['E' + str(total_count)] = "done"
                except:
                    self.newtp_write_sheet['C' + str(total_count)] = "No Contacts"
                    total_count += 1
                    self.driver.close()
                    continue
                # if 'No records to show' in html:
                #     self.newtpsheet['C' + str(total_count)] = "No Contacts"
                #     total_count += 1
                #     self.driver.close()
                #     continue
                # else:
                #     ele = self.wait_id("contact__div")
                #     html = ele.get_attribute('outerHTML')
                #     soup = BeautifulSoup(html, 'html5lib')
                #     tables = soup.findAll('table')
                #     tab = tables[0]
                #     table_body = tab.tbody
                #     tr_group = table_body.find_all('tr')
                #     td_group = tr_group[0].find_all('td')
                #     target = int(len(td_group))
                #     length = int(len(tr_group))
                #     print(target)
                #     print(length)
                #     if target == 1:
                #         self.newtpsheet['C'+str(total_count)] = "No Contacts"
                #     else:
                #         self.newtpsheet['C'+str(total_count)] = tr_group[0].find_all('td')[5].find_all('a')[2].text
                #         if length > 1:
                #         self.newtpsheet['D' + str(total_count)] = tr_group[1].find_all('td')[5].find_all('a')[2].text
                total_count += 1
                self.driver.close()
            else:
                break
        self.info_collect_list_write.save(datetime.now().strftime("%b_%d_%Y") + 'pytest.xlsx')

    def wait(self, xpath):
        try:
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
        finally:
            ele = self.driver.find_element(By.XPATH, xpath)
            return ele

    def wait_id(self, id):
        try:
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, id))
            )
        finally:
            ele = self.driver.find_element(By.ID, id)
            return ele

    def wait_class(self, id):
        try:
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, id))
            )
        finally:
            ele = self.driver.find_element(By.CLASS_NAME, id)
            return ele

    def new_window(self, target_url):
        windows = self.driver.window_handles  # get all handles
        self.driver.switch_to.window(windows[-1])
        self.driver.implicitly_wait(5)
        js = 'window.open("' + target_url + '")'
        self.driver.execute_script(js)
        windows = self.driver.window_handles  # get all handles
        self.driver.switch_to.window(windows[-1])
